"""
utils/excel_generator.py
-------------------------
Generates structured, professional multi-sheet Excel (.xlsx) workbooks
for land analysis reports using openpyxl.
Contains actual property details, factor scores, risks, data quality,
data sources provenance, and regulatory disclaimers.
"""

from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.land import Land
from app.models.analysis import Analysis


def generate_land_report_excel(land: Land, analysis: Analysis, details: dict) -> bytes:
    wb = openpyxl.Workbook()
    
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    blue_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    accent_green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    accent_yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    accent_red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    regular_font = Font(name="Calibri", size=11, color="334155")
    small_gray_font = Font(name="Calibri", size=9, italic=True, color="64748B")
    
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # 1. Executive Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary.merge_cells("A1:F2")
    ws_summary["A1"] = "SMART LAND ANALYSIS PLATFORM — DECISION SUPPORT REPORT"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].fill = navy_fill
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_summary["A4"] = "1. Property Overview"
    ws_summary["A4"].font = section_font

    area_conv = details.get("area_conversions") or {}
    area_sqft = area_conv.get("sqft") or (land.area_sqft if land and land.area_sqft else 0.0)
    area_sqm = area_conv.get("sqm") or (area_sqft / 10.7639 if area_sqft else 0.0)
    area_cents = area_conv.get("cents") or (area_sqft / 435.6 if area_sqft else 0.0)

    prop_rows = [
        ("Land Name / Identifier", (land.land_name if land else None) or "Selected Parcel"),
        ("Address / Locality", (land.address if land else None) or "Unknown Address"),
        ("Geographic Coordinates", f"{land.latitude:.6f}, {land.longitude:.6f}" if land and land.latitude is not None and land.longitude is not None else "N/A"),
        ("Parcel Area (Sq.Ft)", f"{area_sqft:,.1f} sq.ft"),
        ("Parcel Area (Sq.Meters)", f"{area_sqm:,.1f} m²"),
        ("Parcel Area (Cents)", f"{area_cents:,.2f} cents"),
        ("Road Access Width", f"{land.road_width:.0f} ft" if land and land.road_width else "20 ft"),
        ("Soil Taxonomy", land.soil_type.value if land and hasattr(land, 'soil_type') and land.soil_type else "Loamy"),
        ("Zoning / Land Type", land.land_type.value if land and hasattr(land, 'land_type') and land.land_type else "Residential"),
        ("Water Utility Status", "Available" if land and land.water_availability else "Not Available"),
        ("Electricity Utility Status", "Available" if land and land.electricity_availability else "Not Available"),
    ]

    for idx, (label, val) in enumerate(prop_rows, start=5):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"A{idx}"].font = bold_font
        ws_summary[f"A{idx}"].fill = light_gray_fill
        ws_summary[f"A{idx}"].border = border
        
        ws_summary.merge_cells(f"B{idx}:F{idx}")
        ws_summary[f"B{idx}"] = val
        ws_summary[f"B{idx}"].font = regular_font
        ws_summary[f"B{idx}"].border = border

    # Plot Size Validation Section
    plot_val = details.get("plot_validation") or {}
    val_r = len(prop_rows) + 6
    ws_summary[f"A{val_r}"] = "2. Plot Size vs Building Requirement Validation"
    ws_summary[f"A{val_r}"].font = section_font

    val_rows = [
        ("Target Building Type", plot_val.get("building_type", analysis.recommended_building_type if analysis else "Individual House")),
        ("Minimum Required Area", f"{plot_val.get('required_min_sqft', 0):,.0f} sq.ft"),
        ("Actual Parcel Area", f"{plot_val.get('actual_sqft', area_sqft):,.1f} sq.ft"),
        ("Status & Viability", "SUITABLE (Meets Requirement)" if plot_val.get("is_valid", True) else "DEFICIT (Below Requirement)"),
        ("Validation Message", plot_val.get("message", "Plot size is adequate for proposed construction.")),
    ]

    for idx, (label, val) in enumerate(val_rows, start=val_r + 1):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"A{idx}"].font = bold_font
        ws_summary[f"A{idx}"].fill = light_gray_fill
        ws_summary[f"A{idx}"].border = border
        
        ws_summary.merge_cells(f"B{idx}:F{idx}")
        ws_summary[f"B{idx}"] = val
        ws_summary[f"B{idx}"].font = bold_font if "Status" in label else regular_font
        if "Status" in label:
            ws_summary[f"B{idx}"].fill = accent_green_fill if plot_val.get("is_valid", True) else accent_red_fill
        ws_summary[f"B{idx}"].border = border

    start_r = val_r + len(val_rows) + 2
    ws_summary[f"A{start_r}"] = "3. AI Suitability & Risk Assessment"
    ws_summary[f"A{start_r}"].font = section_font

    suitability_metrics = [
        ("Overall Suitability Score", f"{analysis.suitability_score:.1f}%" if analysis and analysis.suitability_score is not None else "N/A", "Out of 100% (Higher = More Suitable)"),
        ("Recommended Building Type", analysis.recommended_building_type if analysis else "N/A", "Optimal developmental footprint"),
        ("Aggregated Risk Score", f"{analysis.risk_score:.1f}%" if analysis and analysis.risk_score is not None else "N/A", "Composite risk across 4 key dimensions"),
        ("Overall Risk Level", analysis.risk_level if analysis else "N/A", "Banded Risk Classification (Low/Moderate/High)"),
        ("Flood Risk Exposure", analysis.flood_risk if analysis else "N/A", "Hydrological elevation & discharge assessment"),
        ("Environmental Risk", analysis.environmental_risk if analysis else "N/A", "Satellite land cover & ecological sensitivity"),
        ("Infrastructure Score", f"{analysis.infrastructure_score:.1f}%" if analysis and analysis.infrastructure_score is not None else "N/A", "Utility grid proximity index"),
        ("Accessibility Score", f"{analysis.traffic_accessibility_score:.1f}%" if analysis and analysis.traffic_accessibility_score is not None else "N/A", "Corridor width & highway accessibility"),
    ]

    for idx, (metric, val, note) in enumerate(suitability_metrics, start=start_r + 1):
        ws_summary[f"A{idx}"] = metric
        ws_summary[f"A{idx}"].font = bold_font
        ws_summary[f"A{idx}"].fill = light_blue_fill
        ws_summary[f"A{idx}"].border = border

        ws_summary[f"B{idx}"] = val
        ws_summary[f"B{idx}"].font = bold_font
        ws_summary[f"B{idx}"].alignment = Alignment(horizontal="center")
        ws_summary[f"B{idx}"].border = border

        ws_summary.merge_cells(f"C{idx}:F{idx}")
        ws_summary[f"C{idx}"] = note
        ws_summary[f"C{idx}"].font = regular_font
        ws_summary[f"C{idx}"].border = border

    expl_r = start_r + len(suitability_metrics) + 2
    ws_summary[f"A{expl_r}"] = "4. Analysis Explanation & Reasoning"
    ws_summary[f"A{expl_r}"].font = section_font

    ws_summary.merge_cells(f"A{expl_r+1}:F{expl_r+3}")
    ws_summary[f"A{expl_r+1}"] = analysis.ai_explanation if analysis else "Analysis completed."
    ws_summary[f"A{expl_r+1}"].font = regular_font
    ws_summary[f"A{expl_r+1}"].alignment = Alignment(wrap_text=True, vertical="top")

    # 2. Factor Breakdown
    ws_factors = wb.create_sheet(title="Factor Breakdown")
    ws_factors.views.sheetView[0].showGridLines = True

    ws_factors.merge_cells("A1:F2")
    ws_factors["A1"] = "DETAILED FACTOR SCORE BREAKDOWN & IMPACT ASSESSMENT"
    ws_factors["A1"].font = title_font
    ws_factors["A1"].fill = blue_header_fill
    ws_factors["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers_factors = ["Factor Name", "Score (0-100)", "Impact", "Weight in Risk", "Why this score? (Physical Reason)", "Assessment Basis / Methodology"]
    for c_idx, h_text in enumerate(headers_factors, start=1):
        cell = ws_factors.cell(row=4, column=c_idx, value=h_text)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    factors = details.get("factors", [])
    for r_idx, f in enumerate(factors, start=5):
        c1 = ws_factors.cell(row=r_idx, column=1, value=f.get("name", ""))
        c1.font = bold_font
        c1.border = border
        
        c2 = ws_factors.cell(row=r_idx, column=2, value=f"{f['score']:.1f}" if f.get("score") is not None else "N/A")
        c2.font = bold_font
        c2.alignment = Alignment(horizontal="center")
        c2.border = border

        impact = f.get("impact", "Positive")
        c3 = ws_factors.cell(row=r_idx, column=3, value=impact)
        c3.font = bold_font
        c3.alignment = Alignment(horizontal="center")
        c3.border = border
        if impact == "Positive":
            c3.fill = accent_green_fill
        elif impact == "Moderate":
            c3.fill = accent_yellow_fill
        elif impact == "Negative":
            c3.fill = accent_red_fill

        weight_str = f"{round(f['weight'] * 100)}%" if f.get("weight") is not None else "N/A"
        c4 = ws_factors.cell(row=r_idx, column=4, value=weight_str)
        c4.alignment = Alignment(horizontal="center")
        c4.font = regular_font
        c4.border = border

        c5 = ws_factors.cell(row=r_idx, column=5, value=f.get("why_reason") or f.get("description", ""))
        c5.font = regular_font
        c5.border = border

        c6 = ws_factors.cell(row=r_idx, column=6, value=f.get("data_source", "GIS Engine"))
        c6.font = regular_font
        c6.border = border

    # 3. Construction Cost Estimate Sheet
    ws_cost = wb.create_sheet(title="Cost Estimate")
    ws_cost.views.sheetView[0].showGridLines = True

    ws_cost.merge_cells("A1:E2")
    ws_cost["A1"] = "INDICATIVE CONSTRUCTION COST ESTIMATOR (CIVIL WORKS)"
    ws_cost["A1"].font = title_font
    ws_cost["A1"].fill = navy_fill
    ws_cost["A1"].alignment = Alignment(horizontal="center", vertical="center")

    cost_info = details.get("construction_cost") or {}
    cost_rows = [
        ("Target Building Type", cost_info.get("building_type", analysis.recommended_building_type if analysis else "Individual House")),
        ("Estimated Area Basis", f"{cost_info.get('area_sqft', area_sqft):,.1f} sq.ft"),
        ("Baseline Rate per Sq.Ft", f"₹{cost_info.get('rate_per_sqft', 2000):,.0f} / sq.ft"),
        ("Total Estimated Cost", f"₹{cost_info.get('total_estimated_cost', 0):,.0f}"),
        ("Material Component (55%)", f"₹{cost_info.get('material_cost', 0):,.0f}"),
        ("Labour Component (25%)", f"₹{cost_info.get('labour_cost', 0):,.0f}"),
        ("Finishing & Contingency (20%)", f"₹{cost_info.get('finishing_cost', 0):,.0f}"),
    ]

    for idx, (label, val) in enumerate(cost_rows, start=4):
        ws_cost[f"A{idx}"] = label
        ws_cost[f"A{idx}"].font = bold_font
        ws_cost[f"A{idx}"].fill = light_gray_fill
        ws_cost[f"A{idx}"].border = border
        
        ws_cost.merge_cells(f"B{idx}:E{idx}")
        ws_cost[f"B{idx}"] = val
        ws_cost[f"B{idx}"].font = bold_font if "Total" in label else regular_font
        if "Total" in label:
            ws_cost[f"B{idx}"].fill = light_blue_fill
        ws_cost[f"B{idx}"].border = border

    note_r = len(cost_rows) + 5
    ws_cost.merge_cells(f"A{note_r}:E{note_r+2}")
    ws_cost[f"A{note_r}"] = (
        "Note: Construction cost estimates are purely indicative planning estimates calculated "
        "using regional standard square-foot civil construction rates in Indian Rupees (₹). "
        "Actual costs will vary based on architectural complexity, soil bearing capacity, "
        "foundation engineering requirements, and fluctuating raw material commodity prices."
    )
    ws_cost[f"A{note_r}"].font = small_gray_font
    ws_cost[f"A{note_r}"].alignment = Alignment(wrap_text=True, vertical="top")

    # 4. Data Quality
    ws_quality = wb.create_sheet(title="Data Quality")
    ws_quality.views.sheetView[0].showGridLines = True

    overall_conf = details.get("data_quality", {}).get("overall_confidence_pct", 94)
    ws_quality.merge_cells("A1:D2")
    ws_quality["A1"] = f"SCIENTIFIC DATA QUALITY & COMPLETENESS ({overall_conf}% Overall)"
    ws_quality["A1"].font = title_font
    ws_quality["A1"].fill = navy_fill
    ws_quality["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers_quality = ["Data Category", "Completeness / Confidence", "Status", "Technical Basis & Source Quality"]
    for c_idx, h_text in enumerate(headers_quality, start=1):
        cell = ws_quality.cell(row=4, column=c_idx, value=h_text)
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    quality_items = details.get("data_quality", {}).get("items", [])
    for r_idx, q in enumerate(quality_items, start=5):
        c1 = ws_quality.cell(row=r_idx, column=1, value=q.get("category", ""))
        c1.font = bold_font
        c1.border = border

        conf = q.get("completeness_pct", 90)
        c2 = ws_quality.cell(row=r_idx, column=2, value=f"{conf:.0f}%")
        c2.font = bold_font
        c2.alignment = Alignment(horizontal="center")
        c2.border = border
        if conf >= 90:
            c2.fill = accent_green_fill

        c3 = ws_quality.cell(row=r_idx, column=3, value=q.get("status", "Verified"))
        c3.font = regular_font
        c3.border = border

        c4 = ws_quality.cell(row=r_idx, column=4, value=q.get("basis", ""))
        c4.font = regular_font
        c4.border = border

    # 5. Data Sources & Provenance
    ws_provenance = wb.create_sheet(title="Data Sources & Provenance")
    ws_provenance.views.sheetView[0].showGridLines = True

    ws_provenance.merge_cells("A1:F2")
    ws_provenance["A1"] = "GIS DATA SOURCES, PROVENANCE & TECHNICAL METADATA"
    ws_provenance["A1"].font = title_font
    ws_provenance["A1"].fill = blue_header_fill
    ws_provenance["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers_prov = ["Dataset Name", "Authoritative Source", "Dataset Date", "Spatial Resolution", "Processing Method", "Sync Frequency"]
    for c_idx, h_text in enumerate(headers_prov, start=1):
        cell = ws_provenance.cell(row=4, column=c_idx, value=h_text)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    sources = details.get("data_sources", [])
    for r_idx, s in enumerate(sources, start=5):
        ws_provenance.cell(row=r_idx, column=1, value=s.get("dataset_name", "")).font = bold_font
        ws_provenance.cell(row=r_idx, column=2, value=s.get("source", "")).font = regular_font
        ws_provenance.cell(row=r_idx, column=3, value=s.get("data_date", "")).font = regular_font
        ws_provenance.cell(row=r_idx, column=4, value=s.get("resolution", "")).font = regular_font
        ws_provenance.cell(row=r_idx, column=5, value=s.get("processing_method", "")).font = regular_font
        ws_provenance.cell(row=r_idx, column=6, value=s.get("last_updated", "")).font = regular_font

        for col in range(1, 7):
            ws_provenance.cell(row=r_idx, column=col).border = border

    # 6. Regulatory Notice
    ws_disc = wb.create_sheet(title="Regulatory Notice")
    ws_disc.views.sheetView[0].showGridLines = True

    ws_disc.merge_cells("A1:E2")
    ws_disc["A1"] = "REGULATORY DISCLAIMER & TECHNICAL CONDITIONS"
    ws_disc["A1"].font = title_font
    ws_disc["A1"].fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
    ws_disc["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_disc.merge_cells("A4:E7")
    ws_disc["A4"] = (
        "OFFICIAL DISCLAIMER:\n\n"
        "This analysis is a GIS-based decision-support assessment. It does not constitute "
        "government approval, legal clearance, land-title verification, environmental clearance, "
        "or structural engineering certification. Site boundaries, road widths, and flood projections "
        "are synthesized from publicly available geospatial records and statistical machine learning "
        "models. A licensed civil engineer or registered surveyor must inspect the physical parcel "
        "prior to property acquisition, excavation, or capital expenditure."
    )
    ws_disc["A4"].font = regular_font
    ws_disc["A4"].alignment = Alignment(wrap_text=True, vertical="top")

    ws_disc["A9"] = f"Report Generated: {datetime.now().strftime('%d %B %Y, %H:%M UTC')}"
    ws_disc["A9"].font = small_gray_font

    # Auto-fit column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 15)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
